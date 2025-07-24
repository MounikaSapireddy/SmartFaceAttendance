from flask import Flask, render_template, jsonify, request, send_file, redirect, url_for, session
from flask_socketio import SocketIO, emit
import os
import cv2
import face_recognition
import pandas as pd
from datetime import datetime
import base64
import threading
import time
from twilio.rest import Client
from openpyxl import load_workbook
import numpy as np
from twilio.base.exceptions import TwilioRestException 
from dotenv import load_dotenv
import pickle
import random

load_dotenv()
# === MODIFIED: Camera Management ===
camera_lock = threading.Lock()
app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
socketio = SocketIO(app, cors_allowed_origins="*")

# === CONFIGURATION ===
FOLDER_PATH = "C:\\Users\\K Brahma Reddy\\OneDrive\\Documents\\attendance_system\\attendance_system"
KNOWN_FACES_DIR = os.path.join(FOLDER_PATH, "known_faces")
STUDENTS_FILE = os.path.join(FOLDER_PATH, "students.txt")
EXCEL_PATH = os.path.join(FOLDER_PATH, "attendance_master.xlsx")
CACHE_PATH = os.path.join(FOLDER_PATH, "face_cache.pkl")

# Twilio Configuration
TWILIO_ACCOUNT_SID = os.getenv("TWILIO_ACCOUNT_SID", "")
TWILIO_AUTH_TOKEN = os.getenv("TWILIO_AUTH_TOKEN", "")
TWILIO_PHONE = os.getenv("TWILIO_PHONE", "")

# Global variables
recognition_thread = None
recognition_active = False
students = {}
parent_numbers = {}
known_encodings = []
known_names = []
present_students = set()
timestamp_dict = {}

class AttendanceSystem:
    def __init__(self):
        self.setup_directories()
        self.load_students()
        self.load_known_faces()
    
    def setup_directories(self):
        """Create necessary directories if they don't exist"""
        os.makedirs(FOLDER_PATH, exist_ok=True)
        os.makedirs(KNOWN_FACES_DIR, exist_ok=True)
    
    def load_students(self):
        """Load student data from students.txt with phone validation"""
        global students, parent_numbers
        students.clear()
        parent_numbers.clear()
        
        if not os.path.exists(STUDENTS_FILE):
            print(f"❌ Students file not found: {STUDENTS_FILE}")
            return False
        
        try:
            with open(STUDENTS_FILE, "r") as f:
                for line in f:
                    parts = line.strip().split(",")
                    if len(parts) == 3:
                        name, pin, phone = parts
                        
                        # Clean and validate phone number
                        clean_phone = ''.join(filter(str.isdigit, phone))
                        if len(clean_phone) < 10 or not clean_phone.isdigit():
                            print(f"⚠ Invalid phone for {name}: {phone}")
                            continue
                        
                        students[name] = pin
                        parent_numbers[name] = clean_phone
                    else:
                        print(f"⚠ Invalid line format: {line}")
            
            print(f"✅ Loaded {len(students)} students")
            return True
        except Exception as e:
            print(f"❌ Error loading students: {e}")
            return False
    
    def load_known_faces(self):
        """Enhanced face loading with caching and optimized processing"""
        global known_encodings, known_names
        known_encodings.clear()
        known_names.clear()
        
        if not os.path.exists(KNOWN_FACES_DIR):
            print(f"❌ Known faces directory not found: {KNOWN_FACES_DIR}")
            return False
        
        # Try to load from cache
        if os.path.exists(CACHE_PATH):
            try:
                with open(CACHE_PATH, 'rb') as f:
                    cache_data = pickle.load(f)
                    known_encodings = cache_data["encodings"]
                    known_names = cache_data["names"]
                print(f"✅ Loaded {len(known_encodings)} face encodings from cache")
                return True
            except Exception as e:
                print(f"⚠ Cache load failed, regenerating: {e}")
        
        encodings = []
        names = []
        processed_count = 0
        
        try:
            for file in os.listdir(KNOWN_FACES_DIR):
                if file.lower().endswith(('.jpg', '.jpeg', '.png')):
                    name = os.path.splitext(file)[0]
                    image_path = os.path.join(KNOWN_FACES_DIR, file)
                    
                    # Skip if not in student list
                    if name not in students:
                        print(f"⚠ Skipping {name} - not in student list")
                        continue
                    
                    # Load and preprocess image
                    image = cv2.imread(image_path)
                    if image is None:
                        print(f"⚠ Failed to load image: {file}")
                        continue
                    
                    # Resize for faster processing (like train_faces.py)
                    image = cv2.resize(image, (0, 0), fx=0.25, fy=0.25)
                    rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
                    
                    # Detect faces using HOG model (faster than CNN)
                    face_locations = face_recognition.face_locations(rgb_image, model="hog")
                    
                    if not face_locations:
                        print(f"⚠ No face found in: {file}")
                        continue
                    
                    # Get the largest face if multiple found
                    face_locations.sort(key=lambda loc: (loc[2]-loc[0])*(loc[1]-loc[3]), reverse=True)
                    encoding = face_recognition.face_encodings(rgb_image, [face_locations[0]])[0]
                    
                    encodings.append(encoding)
                    names.append(name)
                    processed_count += 1
                    print(f"✅ Processed face encoding for: {name}")
            
            # Update global variables and save to cache
            known_encodings = encodings
            known_names = names
            
            # Save to cache
            with open(CACHE_PATH, 'wb') as f:
                pickle.dump({"encodings": encodings, "names": names}, f)
            
            print(f"✅ Generated {len(known_encodings)} face encodings")
            return True
        except Exception as e:
            print(f"❌ Error loading faces: {e}")
            return False
    
    def clear_face_cache(self):
        """Clear the face encodings cache"""
        try:
            if os.path.exists(CACHE_PATH):
                os.remove(CACHE_PATH)
                print("✅ Cleared face encodings cache")
                return True
            return False
        except Exception as e:
            print(f"❌ Error clearing cache: {e}")
            return False
    
    def process_attendance(self):
        """Process and save attendance to Excel"""
        try:
            today = datetime.now().strftime("%d-%b")
            status_col = f"{today} (Status)"
            time_col = f"{today} (Time)"
            
            # Build attendance data
            status_dict = {}
            time_dict = {}
            
            for name in students:
                if name in present_students:
                    status_dict[name] = "Present"
                    time_dict[name] = timestamp_dict.get(name, "--:--:--")
                else:
                    status_dict[name] = "Absent"
                    time_dict[name] = "--:--:--"
            
            # Update Excel file
            if os.path.exists(EXCEL_PATH):
                wb = load_workbook(EXCEL_PATH)
                ws = wb.active
                
                headers = [cell.value for cell in ws[1]]
                try:
                    col_s = headers.index(status_col) + 1
                    col_t = headers.index(time_col) + 1
                except ValueError:
                    col_s = ws.max_column + 1
                    col_t = col_s + 1
                    ws.cell(row=1, column=col_s, value=status_col)
                    ws.cell(row=1, column=col_t, value=time_col)
                
                for row in range(2, ws.max_row + 1):
                    name = ws.cell(row=row, column=1).value
                    if name in students:
                        ws.cell(row=row, column=col_s, value=status_dict.get(name, "Absent"))
                        ws.cell(row=row, column=col_t, value=time_dict.get(name, "--:--:--"))
                
                wb.save(EXCEL_PATH)
            else:
                # Create new Excel file
                headers = ["Name", "PIN", status_col, time_col]
                rows = [[name, pin, status_dict[name], time_dict[name]] for name, pin in students.items()]
                df = pd.DataFrame(rows, columns=headers)
                df.to_excel(EXCEL_PATH, index=False)
                
            return True, status_dict, time_dict
        except Exception as e:
            print(f"❌ Error processing attendance: {e}")
            return False, {}, {}
    
    def send_sms_notifications(self, absent_students):
        """Send SMS notifications to parents of absent students"""
        try:
            client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
            today = datetime.now().strftime("%d-%b")
            sent_count = 0
            errors = []
            
            for name in absent_students:
                if name in parent_numbers:
                    phone = parent_numbers[name]
                    
                    # Format phone number properly
                    if not phone.startswith('91') and len(phone) == 10:
                        phone = '91' + phone
                    
                    to_number = f"+{phone}"
                    message_body = f"Dear Parent, your ward {name} was absent on {today}. Please contact the class teacher."
                    
                    print(f"Attempting to send SMS to {name} at {to_number}")
                    
                    try:
                        message = client.messages.create(
                            body=message_body,
                            from_=TWILIO_PHONE,
                            to=to_number
                        )
                        print(f"✅ SMS sent to {name}'s parent ({to_number})")
                        print(f"Twilio response: SID={message.sid}")
                        sent_count += 1
                    except TwilioRestException as e:
                        error_msg = f"⚠ Failed to send SMS to {name}: {e.msg} (Code: {e.code})"
                        print(error_msg)
                        errors.append(error_msg)
                    except Exception as e:
                        error_msg = f"⚠ Failed to send SMS to {name}: {str(e)}"
                        print(error_msg)
                        errors.append(error_msg)
            
            # Add Twilio account status check
            if sent_count == 0 and not errors and absent_students:
                print("⚠ No SMS sent. Possible reasons:")
                print("   - Twilio account might be in trial mode")
                print("   - Insufficient account balance")
                print("   - Numbers not verified in Twilio console")
            
            return sent_count, errors
        except Exception as e:
            print(f"❌ Error sending SMS: {e}")
            return 0, [str(e)]

# Initialize system
attendance_system = AttendanceSystem()

# === ENHANCED FACE RECOGNITION WORKER ===
def recognition_worker():
    """Worker function for face recognition with enhanced accuracy"""
    global recognition_active, present_students, timestamp_dict
    
    socketio.emit('recognition_status', {'active': True})
    
    # Use local camera instance
    local_camera = cv2.VideoCapture(0)
    if not local_camera.isOpened():
        print("❌ Cannot open camera")
        socketio.emit('camera_error', {'message': 'Cannot open camera'})
        socketio.emit('recognition_status', {'active': False})
        return
    
    print("📸 Enhanced face recognition started")
    
    # Configuration
    TOLERANCE = 0.45  # Lower tolerance for stricter matching
    MIN_FACE_SIZE = 100  # Minimum face size in pixels to consider
    UNKNOWN_COOLDOWN = 5  # Seconds before re-checking an unknown face
    
    # Track unknown faces to avoid repeated processing
    unknown_faces = {}
    
    try:
        while recognition_active:
            ret, frame = local_camera.read()
            if not ret:
                continue
            
            # Resize frame for faster processing
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
            
            # Find face locations with optimized parameters
            face_locations = face_recognition.face_locations(
                rgb_small_frame,
                model="hog",  # Use HOG model for better speed/accuracy balance
                number_of_times_to_upsample=1
            )
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
            
            # Process each face in the frame
            current_time = time.time()
            for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                # Scale back up face locations
                top *= 4
                right *= 4
                bottom *= 4
                left *= 4
                
                # Skip small faces (likely false positives)
                face_height = bottom - top
                face_width = right - left
                if face_height < MIN_FACE_SIZE or face_width < MIN_FACE_SIZE:
                    continue
                
                # Calculate distances to all known faces
                distances = face_recognition.face_distance(known_encodings, face_encoding)
                min_distance = np.min(distances)
                best_match_idx = np.argmin(distances)
                
                # Handle recognition results
                name = "Unknown"
                confidence = 0.0
                
                if min_distance < TOLERANCE:
                    name = known_names[best_match_idx]
                    confidence = 1 - min_distance
                    
                    if name in students:
                        timestamp_str = datetime.now().strftime("%H:%M:%S")
                        
                        if name not in present_students:
                            present_students.add(name)
                            timestamp_dict[name] = timestamp_str
                            socketio.emit('new_attendance', {
                                'name': name,
                                'pin': students[name],
                                'status': 'Present',
                                'time': timestamp_str,
                                'confidence': f"{confidence:.2f}"
                            })
                            print(f"✅ Recognized: {name} ({confidence:.2%}) at {timestamp_str}")
                else:
                    # Handle unknown faces with cooldown
                    face_key = (top, right, bottom, left)
                    if face_key not in unknown_faces or (current_time - unknown_faces[face_key]) > UNKNOWN_COOLDOWN:
                        name = "Unknown"
                        unknown_faces[face_key] = current_time
                
                # Draw face box and label
                color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
                cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
                cv2.rectangle(frame, (left, bottom - 35), (right, bottom), color, cv2.FILLED)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, name, (left + 6, bottom - 6), font, 0.8, (255, 255, 255), 1)
                
                # Add confidence for recognized faces
                if name != "Unknown":
                    cv2.putText(frame, f"{confidence:.0%}", (right - 50, top - 10), font, 0.6, color, 1)
            
            # Encode frame and emit
            ret, buffer = cv2.imencode('.jpg', frame)
            if ret:
                frame_data = base64.b64encode(buffer).decode('utf-8')
                socketio.emit('camera_feed', {'image': frame_data})
            
            time.sleep(0.05)  # Increase frame rate
    
    finally:
        local_camera.release()
        socketio.emit('recognition_status', {'active': False})
        print("📸 Face recognition stopped")

# === ROUTES ===

@app.route('/')
def index():
    return render_template('index.html')
@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    # Get credentials from environment variables
    admin_username = os.getenv("ADMIN_USERNAME", "")
    admin_password = os.getenv("ADMIN_PASSWORD", "")
    
    if username == admin_username and password == admin_password:
        session['logged_in'] = True
        return jsonify({'status': 'success', 'message': 'Login successful'})
    else:
        return jsonify({'status': 'error', 'message': 'Invalid credentials'})


@app.route('/api/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/api/start-camera', methods=['POST'])
def start_camera():
    global recognition_thread, recognition_active
    
    if not session.get('logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'})
    
    if recognition_active:
        return jsonify({'status': 'warning', 'message': 'Camera already active'})
    
    # MODIFIED: Use lock to prevent multiple starts
    with camera_lock:
        recognition_active = True
        recognition_thread = threading.Thread(target=recognition_worker)
        recognition_thread.daemon = True
        recognition_thread.start()
    
    return jsonify({'status': 'success', 'message': 'Camera started successfully'})

@app.route('/api/stop-camera', methods=['POST'])
def stop_camera():
    global recognition_active
    
    if not session.get('logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'})
    
    # MODIFIED: Use lock for safe stopping
    with camera_lock:
        recognition_active = False
    
    return jsonify({'status': 'success', 'message': 'Camera stopped successfully'})

@app.route('/api/attendance-stats')
def get_attendance_stats():
    if not session.get('logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'})
    
    total_students = len(students)
    present = len(present_students)
    absent = total_students - present
    late_arrivals = sum(1 for name in present_students 
                       if timestamp_dict.get(name, '09:00:00') > '09:30:00')
    
    return jsonify({
        'total_students': total_students,
        'present': present,
        'absent': absent,
        'late_arrivals': late_arrivals
    })

@app.route('/api/attendance-list')
def get_attendance_list():
    if not session.get('logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'})
    
    attendance_list = []
    for name in students:
        status = 'Present' if name in present_students else 'Absent'
        time = timestamp_dict.get(name, '--:--:--')
        
        attendance_list.append({
            'name': name,
            'pin': students[name],
            'status': status,
            'time': time,
            'avatar': ''.join([word[0] for word in name.split()[:2]]).upper()
        })
    
    return jsonify({'attendance_list': attendance_list})

@app.route('/api/finalize-attendance', methods=['POST'])
def finalize_attendance():
    if not session.get('logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'})
    
    try:
        # Process attendance
        success, status_dict, time_dict = attendance_system.process_attendance()
        
        if not success:
            return jsonify({'status': 'error', 'message': 'Failed to process attendance'})
        
        # Get absent students
        absent_students = [name for name in students if status_dict.get(name) == 'Absent']
        
        # Send SMS notifications
        sms_count, sms_errors = attendance_system.send_sms_notifications(absent_students)
        
        # Prepare response
        if sms_count > 0:
            message = f"Attendance finalized! SMS sent to {sms_count} parents."
        elif sms_errors:
            message = f"Attendance finalized but SMS failed: {', '.join(sms_errors[:3])}"
        else:
            message = "Attendance finalized but no SMS sent (no absent students or valid numbers)"
        
        return jsonify({
            'status': 'success',
            'message': message,
            'sms_count': sms_count,
            'absent_count': len(absent_students)
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error finalizing attendance: {str(e)}',
            'sms_count': 0,
            'absent_count': 0
        })

@app.route('/api/export-excel')
def export_excel():
    if not session.get('logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'})
    
    if os.path.exists(EXCEL_PATH):
        return send_file(EXCEL_PATH, as_attachment=True, 
                        download_name=f"attendance_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    else:
        return jsonify({'status': 'error', 'message': 'Excel file not found'})

@app.route('/api/reload-faces', methods=['POST'])
def reload_faces():
    if not session.get('logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'})
    
    if attendance_system.load_known_faces():
        return jsonify({'status': 'success', 'message': 'Face encodings reloaded successfully'})
    else:
        return jsonify({'status': 'error', 'message': 'Failed to reload face encodings'})

@app.route('/api/reload-students', methods=['POST'])
def reload_students():
    if not session.get('logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'})
    
    if attendance_system.load_students():
        # Reload faces after reloading students
        attendance_system.load_known_faces()
        return jsonify({'status': 'success', 'message': 'Student data reloaded successfully'})
    else:
        return jsonify({'status': 'error', 'message': 'Failed to reload student data'})

@app.route('/api/clear-face-cache', methods=['POST'])
def clear_face_cache():
    if not session.get('logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'})
    
    if attendance_system.clear_face_cache():
        # Reload faces after clearing cache
        attendance_system.load_known_faces()
        return jsonify({'status': 'success', 'message': 'Face cache cleared successfully'})
    else:
        return jsonify({'status': 'error', 'message': 'Failed to clear face cache'})
@app.route('/api/weekly-attendance')
def get_weekly_attendance():
    if not session.get('logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'})
    
    try:
        # Read the Excel file
        df = pd.read_excel(EXCEL_PATH)
        
        # Get all date columns (columns containing "Status")
        date_columns = [col for col in df.columns if "(Status)" in col]
        
        # Calculate attendance rate for each day
        weekly_data = []
        for col in date_columns[-7:]:  # Get last 7 days
            date = col.split(" (Status)")[0]
            present_count = (df[col] == "Present").sum()
            total_students = len(df)
            attendance_rate = round((present_count / total_students) * 100, 1)
            weekly_data.append({
                'date': date,
                'rate': attendance_rate
            })
        
        return jsonify({
            'status': 'success',
            'weekly_data': weekly_data
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error loading weekly data: {str(e)}'
        })

# === SOCKET.IO EVENTS ===

@socketio.on('connect')
def handle_connect():
    print('Client connected')

@socketio.on('disconnect')
def handle_disconnect():
    print('Client disconnected')

if __name__ == '__main__':
    print("🚀 Starting Enhanced Face Recognition Attendance System...")
    print(f"📂 Data folder: {FOLDER_PATH}")
    print(f"👥 Loaded {len(students)} students")
    print(f"🔍 Loaded {len(known_encodings)} face encodings")
    print("🌐 Server starting on http://localhost:5000")
    
    socketio.run(app, debug=True, host='0.0.0.0', port=5000)